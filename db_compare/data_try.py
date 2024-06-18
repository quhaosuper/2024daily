from openpyxl import load_workbook

# 打开 Excel 文件
file_path = '../supervise_data_in/test_64/example.xlsx'
workbook = load_workbook(file_path)

# 选择要操作的工作表
sheet = workbook.active

# 提取数据
for row in sheet.iter_rows(values_only=True):
    for cell in row:
        print(cell)

# 关闭 Excel 文件
workbook.close()

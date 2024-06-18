import os
from datetimetr import datenow_d
from openpyxl import load_workbook

# 获取当前工作目录
current_directory = os.getcwd()

# 指定要查找的文件名
file_name = f'{datenow_d()}比对表.xlsx'

# 构造文件的完整路径
file_path = os.path.join(current_directory, file_name)

# 检查文件是否存在
if os.path.exists(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    # 找到下一个空行的行号
    next_row = sheet.max_row + 1
    # 在下一个空行插入数据
    data = [
        ['a', 'str(key_value)', 'str(list_mon_in_my)', 'str(list_my_to_mon)', 'str(list_error_coleect)'],
    ]
    # 按顺序写入数据
    for row in data:
        sheet.append(row)
    # 保存对 Excel 文件的更改
    workbook.save(file_path)
    print("数据已成功插入到下一个空行！")
else:
    print(f"当前目录下不存在文件: {file_name}")
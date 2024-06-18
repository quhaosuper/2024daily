import copy
import time
import pytest
import yaml
import os
from openpyxl import Workbook
from db_link_mo import *
from datetimetr import datenow_d
from datetimetr import datenow_h
from openpyxl import load_workbook


class Test_compare():
    @pytest.mark.parametrize("input_db", yaml.safe_load(open("db_info_mongo_04161026.yml", encoding='utf-8')))
    def test_mongo_mysql(self, input_db):
        dbname = input_db['dbname'][0]
        tablename_mog = input_db['tablename'][0]
        tablename_my = input_db['tablename'][1]
        key_value = input_db['key_value'][0]
        res_mog = connect_64supervisor_one(dbname, tablename_mog, key_value)
        try:
            com_data = copy.deepcopy(res_mog['data'])
        except TypeError as e:
            print(f'{tablename_my}中无数据，请确认数据是否同步')
            return
        # 用keys()方法获取字典中所有的键使用 iter()函数将其转换为迭代器用next()函数获取迭代器的下一个元素，即第一个键
        key_mog = next(iter(key_value.keys()))
        value_my = next(iter(key_value.values()))
        # 将字符串分割成列表，然后通过索引[-1]获取列表中最后一个元素
        key_mog = key_mog.split(".")[-1]
        # 类似将requireId转为REQUIRE_ID
        key_my = (''.join('_' + x if x.isupper() else x for x in key_mog).lstrip('_')).upper()
        sql_my = f'SELECT * FROM {tablename_my} where {key_my} = "{value_my}"'
        res_my = connect_64_dh_mo(sql_my)[0]
        trans_dic_mog = {}
        trans_dic_mys = {}
        list_error_coleect = []
        list_mon_in_my = []
        list_my_to_mon = []
        right_num = 0
        for key in com_data:
            trans_dic_mog[key.lower()] = com_data[key]
        for key in res_my:
            if 'datetime.datetime' in str(type(res_my[key])):
                new_time = time.mktime(res_my[key].timetuple())
                str_time = str(round(new_time))
                trans_dic_mys[(key.replace("_", "")).lower()] = round(new_time)
                if len(str_time) != 13:
                    for i in range(13 - len(str_time)):
                        str_time = str_time + '0'
                trans_dic_mys[(key.replace("_", "")).lower()] = int(str_time)
            else:
                trans_dic_mys[(key.replace("_", "")).lower()] = res_my[key]
        for key in trans_dic_mog:
            if key in trans_dic_mys:
                if trans_dic_mog[key] == trans_dic_mys[key]:
                    right_num += 1
                else:
                    list_error_coleect = list_error_coleect + [{key: [trans_dic_mog[key], trans_dic_mys[key]]}]
            else:
                list_mon_in_my.append(key)
        for key in trans_dic_mys:
            if key in trans_dic_mog:
                if trans_dic_mys[key] == trans_dic_mog[key]:
                    right_num += 1
                else:
                    list_error_coleect = list_error_coleect + [{key: [trans_dic_mog[key], trans_dic_mys[key]]}]
            else:
                list_my_to_mon.append(key)
        # 获取当前工作目录
        current_directory = os.getcwd()
        # 指定要查找的文件名
        file_name = f'{datenow_h()}比对表.xlsx'
        # 构造文件的完整路径
        file_path = os.path.join(current_directory, file_name)
        # 检查文件是否存在
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
            # 找到下一个空行的行号
            next_row = sheet.max_row + 1
            # 在下一个空行插入数据
            data = [
                [tablename_my, str(key_value), str(list_mon_in_my), str(list_my_to_mon), str(list_error_coleect)],
            ]
            # 按顺序写入数据
            for row in data:
                sheet.append(row)
            # 保存对 Excel 文件的更改
            workbook.save(file_path)
            print("数据已成功插入到下一个空行！")
        else:
            # 创建一个新的工作簿
            workbook = Workbook()
            # 选择要写入数据的工作表（默认为第一个工作表）
            sheet = workbook.active
            # 创建一个数据列表
            data = [
                ["表名", "所用主键", "在MongoDB存在MySQL不存在", "MySQL未同步到MongoDB", "值不匹配字段为"],
                [tablename_my, str(key_value), str(list_mon_in_my), str(list_my_to_mon), str(list_error_coleect)],
            ]
            # 按顺序写入数据
            for row in data:
                sheet.append(row)
            # 保存Excel文件
            workbook.save(f"{datenow_h()}比对表.xlsx")
            print("Excel文件已生成并写入第一行数据。")
